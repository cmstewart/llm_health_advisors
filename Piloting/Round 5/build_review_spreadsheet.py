#!/usr/bin/env python3
"""
Build a review spreadsheet from the merged analysis dataset.

One tab per opening post, containing the OP, the real clinician comments, and all
nine model x strategy synthetic sets stacked underneath for side-by-side reading.
An index tab up front summarises the selection.

Usage:
    python build_review_spreadsheet.py                  # 10 OPs
    python build_review_spreadsheet.py --n-ops 25
    python build_review_spreadsheet.py --n-ops 10 --seed 7
"""

from __future__ import annotations

import argparse
import json
import random
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

MODELS = ["gemini", "openai", "grok"]
STRATEGIES = ["dg", "sag", "sagii"]

# One fill per source, so a reader can tell the blocks apart at a glance
FILLS = {
    "real": "FCE4D6",    # warm
    "gemini": "E2EFDA",  # green
    "openai": "DDEBF7",  # blue
    "grok": "EDE7F6",    # purple
}

LABELS = {
    "dg": "DG (whole thread in one prompt)",
    "sag": "SAG (independent, sees only the OP)",
    "sagii": "SAGII (sequential, sees prior comments)",
}

ILLEGAL_SHEET_CHARS = r'[\[\]\*\?/\\:]'


def sheet_name(idx: int, title: str) -> str:
    """Excel: max 31 chars, and none of [ ] * ? / \\ :"""
    clean = re.sub(ILLEGAL_SHEET_CHARS, "", title)
    name = f"{idx}. {clean}"
    return name[:31]


def select(records: list[dict], n: int, seed: int) -> list[dict]:
    """
    Pick a readable spread rather than a pure random draw: prefer threads with
    2-6 real comments (enough to compare, not so many the tab is unreadable),
    and favour ones where a real comment asked a question, since that is the
    pattern the analysis turns on.
    """
    rng = random.Random(seed)
    mid = [r for r in records if 2 <= len(r["real_comments"]) <= 6]
    pool = mid or records

    asked = [r for r in pool if any("?" in c["body"] for c in r["real_comments"])]
    quiet = [r for r in pool if r not in asked]

    n_asked = min(len(asked), round(n * 0.6))
    picked = rng.sample(asked, n_asked) + rng.sample(quiet, min(len(quiet), n - n_asked))
    rng.shuffle(picked)
    return picked[:n]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default="../output/corpora/analysis_dataset.jsonl")
    ap.add_argument("--out", default="../output/corpora/review_sample.xlsx")
    ap.add_argument("--n-ops", type=int, default=10)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    records = [json.loads(line) for line in open(args.data)]
    print(f"Loaded {len(records):,} threads")

    chosen = select(records, args.n_ops, args.seed)
    print(f"Selected {len(chosen)} for review")

    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=12, name="Arial")
    header_font = Font(bold=True, size=11, name="Arial")
    body_font = Font(size=10, name="Arial")
    italic = Font(size=9, name="Arial", italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    rule = Border(bottom=Side(style="thin", color="CCCCCC"))

    # ---------- index ----------
    idx = wb.create_sheet("Index")
    idx.column_dimensions["A"].width = 6
    idx.column_dimensions["B"].width = 70
    idx.column_dimensions["C"].width = 14
    idx.column_dimensions["D"].width = 20

    idx["A1"] = "Review sample: real vs. synthetic clinician comments"
    idx["A1"].font = title_font
    idx["A2"] = (
        f"{len(chosen)} threads drawn from {len(records):,} (seed {args.seed}). "
        "Each tab shows the opening post, the real clinician replies, and all nine "
        "model x strategy synthetic sets."
    )
    idx["A2"].font = italic
    idx["A2"].alignment = wrap

    for col, head in zip("ABCD", ["Tab", "Opening post", "Real replies", "Real asked a question"]):
        idx[f"{col}4"] = head
        idx[f"{col}4"].font = header_font

    for i, rec in enumerate(chosen, 1):
        asked = any("?" in c["body"] for c in rec["real_comments"])
        idx.cell(row=4 + i, column=1, value=i).font = body_font
        idx.cell(row=4 + i, column=2, value=rec["title"]).font = body_font
        idx.cell(row=4 + i, column=2).alignment = wrap
        idx.cell(row=4 + i, column=3, value=len(rec["real_comments"])).font = body_font
        idx.cell(row=4 + i, column=4, value="yes" if asked else "no").font = body_font

    # ---------- one tab per OP ----------
    for i, rec in enumerate(chosen, 1):
        ws = wb.create_sheet(sheet_name(i, rec["title"]))
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 100
        row = 1

        def section(label: str, fill_key: str) -> None:
            nonlocal row
            for col in (1, 2):
                c = ws.cell(row=row, column=col, value=label if col == 1 else None)
                c.fill = PatternFill("solid", fgColor=FILLS[fill_key])
                c.font = title_font
            row += 1

        def line(left: str, right: str, note: str = "") -> None:
            nonlocal row
            a = ws.cell(row=row, column=1, value=left)
            b = ws.cell(row=row, column=2, value=right)
            a.font = italic if note else body_font
            b.font = body_font
            a.alignment = wrap
            b.alignment = wrap
            a.border = rule
            b.border = rule
            row += 1

        # opening post
        section("OPENING POST", "real")
        line("Title", rec["title"])
        line("Body", rec["selftext"])
        line("Thread", f"{len(rec['real_comments'])} real repl(y/ies)  |  id {rec['submission_id']}")
        row += 1

        # real comments
        section("REAL COMMENTS", "real")
        line("Flair", "Comment")
        for c in rec["real_comments"]:
            flair = c["author_flair_text"] or "no flair"
            mark = "  [asks a question]" if "?" in c["body"] else ""
            line(flair + mark, c["body"])
        row += 1

        # synthetic
        for m in MODELS:
            for s in STRATEGIES:
                cl = rec["synthetic"][m][s]
                section(f"{m.upper()} / {s.upper()}", m)
                ws.cell(row=row - 1, column=2, value=LABELS[s]).font = italic
                if not cl:
                    line("", "(no output)")
                for c in cl:
                    mark = "[asks a question]" if "?" in c["body"] else ""
                    line(mark, c["body"])
                row += 1

    out = Path(args.out)
    wb.save(out)
    print(f"\nSaved {out}  ({out.stat().st_size / 1e6:.1f} MB)")
    print(f"Tabs: Index + {len(chosen)} threads")


if __name__ == "__main__":
    main()
